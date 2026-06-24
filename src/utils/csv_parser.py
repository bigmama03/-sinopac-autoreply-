"""CSV and Excel template file parser."""

import csv
import io
from typing import Optional

from src.data.models import Template

# Expected columns
REQUIRED_COLUMNS = {"template_code", "category", "content", "platforms"}

# Column name aliases (Traditional Chinese → English)
COLUMN_ALIASES = {
    "文案編號": "template_code",
    "編號": "template_code",
    "分類": "category",
    "類別": "category",
    "文案內容": "content",
    "內容": "content",
    "適用平台": "platforms",
    "平台": "platforms",
}


def _normalize_columns(headers: list[str]) -> dict[str, int]:
    """Map header names to column indices, resolving aliases."""
    mapping = {}
    for idx, header in enumerate(headers):
        h = header.strip()
        if h in REQUIRED_COLUMNS:
            mapping[h] = idx
        elif h in COLUMN_ALIASES:
            mapping[COLUMN_ALIASES[h]] = idx
    return mapping


def _validate_columns(col_map: dict[str, int]) -> Optional[str]:
    """Return error message if required columns are missing."""
    missing = REQUIRED_COLUMNS - col_map.keys()
    if missing:
        return f"缺少必要欄位: {', '.join(missing)}"
    return None


def parse_csv(file_path: str) -> tuple[list[Template], Optional[str]]:
    """Parse a CSV file into templates. Returns (templates, error_message)."""
    templates = []
    try:
        # Try UTF-8 first, then Big5 (common in Taiwan)
        content = None
        for encoding in ("utf-8-sig", "utf-8", "big5", "cp950"):
            try:
                with open(file_path, "r", encoding=encoding) as f:
                    content = f.read()
                break
            except UnicodeDecodeError:
                continue

        if content is None:
            return [], "無法辨識檔案編碼，請使用 UTF-8 或 Big5 編碼"

        reader = csv.reader(io.StringIO(content))
        headers = next(reader, None)
        if not headers:
            return [], "CSV 檔案為空"

        col_map = _normalize_columns(headers)
        error = _validate_columns(col_map)
        if error:
            return [], error

        max_col_idx = max(col_map.values())

        for row_num, row in enumerate(reader, start=2):
            if not any(cell.strip() for cell in row):
                continue  # Skip blank rows

            # Pad short rows to avoid IndexError
            if len(row) <= max_col_idx:
                row = row + [""] * (max_col_idx - len(row) + 1)

            code = row[col_map["template_code"]].strip()
            category = row[col_map["category"]].strip()
            content_text = row[col_map["content"]].strip()
            platforms = row[col_map["platforms"]].strip()

            if not code or not content_text:
                continue  # Skip incomplete rows

            templates.append(Template(
                template_code=code,
                category=category,
                content=content_text,
                platforms=platforms,
            ))

        if not templates:
            return [], "檔案中沒有有效的文案資料"

        return templates, None

    except Exception as e:
        return [], f"解析 CSV 失敗: {str(e)}"


def parse_excel(file_path: str) -> tuple[list[Template], Optional[str]]:
    """Parse an Excel file into templates. Returns (templates, error_message)."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        return [], "缺少 openpyxl 套件，無法讀取 Excel 檔案"

    templates = []
    try:
        wb = load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return [], "Excel 檔案為空"

        headers = [str(cell) if cell else "" for cell in rows[0]]
        col_map = _normalize_columns(headers)
        error = _validate_columns(col_map)
        if error:
            return [], error

        for row in rows[1:]:
            if not any(cell for cell in row):
                continue

            code = str(row[col_map["template_code"]] or "").strip()
            category = str(row[col_map["category"]] or "").strip()
            content_text = str(row[col_map["content"]] or "").strip()
            platforms = str(row[col_map["platforms"]] or "").strip()

            if not code or not content_text:
                continue

            templates.append(Template(
                template_code=code,
                category=category,
                content=content_text,
                platforms=platforms,
            ))

        wb.close()

        if not templates:
            return [], "檔案中沒有有效的文案資料"

        return templates, None

    except Exception as e:
        return [], f"解析 Excel 失敗: {str(e)}"


def parse_file(file_path: str) -> tuple[list[Template], Optional[str]]:
    """Auto-detect file type and parse."""
    lower = file_path.lower()
    if lower.endswith(".csv"):
        return parse_csv(file_path)
    elif lower.endswith((".xlsx", ".xls")):
        return parse_excel(file_path)
    else:
        return [], "不支援的檔案格式，請使用 CSV 或 Excel (.xlsx) 檔案"


KEYWORD_REQUIRED_COLUMNS = {"keyword", "category", "weight"}

KEYWORD_COLUMN_ALIASES = {
    "關鍵字": "keyword",
    "分類": "category",
    "類別": "category",
    "權重": "weight",
}


def parse_keyword_csv(file_path):
    """Parse a keyword CSV file. Returns (list[dict], None) or ([], error_message)."""

    content = None
    for encoding in ("utf-8-sig", "utf-8", "big5", "cp950"):
        try:
            with open(file_path, "r", encoding=encoding) as f:
                content = f.read()
            break
        except UnicodeDecodeError:
            continue

    if content is None:
        return [], "無法辨識檔案編碼，請使用 UTF-8 或 Big5 編碼"

    reader = csv.reader(io.StringIO(content))
    headers = next(reader, None)
    if not headers:
        return [], "CSV 檔案為空"

    # Normalize header names
    col_map = {}
    for idx, header in enumerate(headers):
        h = header.strip()
        if h in KEYWORD_REQUIRED_COLUMNS:
            col_map[h] = idx
        elif h in KEYWORD_COLUMN_ALIASES:
            col_map[KEYWORD_COLUMN_ALIASES[h]] = idx

    missing = KEYWORD_REQUIRED_COLUMNS - col_map.keys()
    if missing:
        return [], f"缺少必要欄位: {', '.join(missing)}"

    max_col_idx = max(col_map.values())
    entries = []
    for row in reader:
        if not any(cell.strip() for cell in row):
            continue
        if len(row) <= max_col_idx:
            row = row + [""] * (max_col_idx - len(row) + 1)

        kw = row[col_map["keyword"]].strip()
        cat = row[col_map["category"]].strip()
        wt_raw = row[col_map["weight"]].strip()
        if not kw or not wt_raw:
            continue
        try:
            wt = float(wt_raw)
        except ValueError:
            continue
        if not (1.0 <= wt <= 5.0):
            continue
        entries.append({"keyword": kw, "category": cat, "weight": wt})

    if not entries:
        return [], "檔案中沒有有效的關鍵字資料"

    return entries, None


def parse_keyword_excel(file_path):
    """Parse a keyword Excel file. Returns (list[dict], None) or ([], error_message)."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        return [], "缺少 openpyxl 套件，無法讀取 Excel 檔案"

    try:
        wb = load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            wb.close()
            return [], "Excel 檔案為空"

        headers_raw = [str(cell).strip() if cell else "" for cell in rows[0]]
        headers = [
            KEYWORD_COLUMN_ALIASES.get(h, h) for h in headers_raw
        ]

        missing = KEYWORD_REQUIRED_COLUMNS - set(headers)
        if missing:
            wb.close()
            return [], f"缺少必要欄位: {', '.join(missing)}"

        entries = []
        for raw_row in rows[1:]:
            if not any(cell for cell in raw_row):
                continue
            row = {
                headers[i]: (str(raw_row[i]).strip() if raw_row[i] is not None else "")
                for i in range(len(headers))
            }
            kw = row.get("keyword", "").strip()
            cat = row.get("category", "").strip()
            wt_raw = row.get("weight", "").strip()
            if not kw or not wt_raw:
                continue
            try:
                wt = float(wt_raw)
            except ValueError:
                continue
            if not (1.0 <= wt <= 5.0):
                continue
            entries.append({"keyword": kw, "category": cat, "weight": wt})

        wb.close()

        if not entries:
            return [], "檔案中沒有有效的關鍵字資料"

        return entries, None

    except Exception as e:
        return [], f"解析 Excel 失敗: {str(e)}"


def parse_keyword_file(file_path):
    """Dispatch to parse_keyword_csv or parse_keyword_excel based on extension."""
    import os
    ext = os.path.splitext(file_path)[1].lower()
    if ext == ".xlsx":
        return parse_keyword_excel(file_path)
    elif ext == ".csv":
        return parse_keyword_csv(file_path)
    else:
        return [], f"不支援的檔案格式：{ext}"
